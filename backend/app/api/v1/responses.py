from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import List
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database import get_db
from app.models import SurveyResponse, Answer, Survey, Question, QuestionOption, Employee
from app.schemas import ResponseList, SurveyResults, ResponseResult, QuestionResult, EmployeeResult, SurveyResponse as SurveyResponseSchema, SurveyAnalytics, QuestionAnalytics

router = APIRouter()


@router.get("", response_model=ResponseList)
async def get_responses(
    skip: int = 0,
    limit: int = 100,
    survey_id: int = None,
    db: AsyncSession = Depends(get_db)
):
    """Get list of all responses."""
    query = select(SurveyResponse).options(selectinload(SurveyResponse.answers))

    if survey_id:
        query = query.where(SurveyResponse.survey_id == survey_id)

    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    responses = result.scalars().all()

    # Get total count
    count_query = select(SurveyResponse)
    if survey_id:
        count_query = count_query.where(SurveyResponse.survey_id == survey_id)
    count_result = await db.execute(count_query)
    total = len(count_result.scalars().all())

    return ResponseList(responses=responses, total=total)


@router.get("/surveys/{survey_id}/results", response_model=SurveyResults)
async def get_survey_results(survey_id: int, db: AsyncSession = Depends(get_db)):
    """Get survey results in JSON format."""
    # Get survey
    survey_result = await db.execute(
        select(Survey)
        .where(Survey.id == survey_id)
        .options(selectinload(Survey.questions).selectinload(Question.options))
    )
    survey = survey_result.scalar_one_or_none()

    if not survey:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Survey not found"
        )

    # Get all responses with answers and employee
    responses_result = await db.execute(
        select(SurveyResponse)
        .where(SurveyResponse.survey_id == survey_id)
        .options(
            selectinload(SurveyResponse.answers),
            selectinload(SurveyResponse.employee)
        )
    )
    responses = responses_result.scalars().all()

    # Get total eligible employees (active employees who started 90+ days ago)
    from datetime import datetime, timedelta
    ninety_days_ago = datetime.now().date() - timedelta(days=survey.days_after_start)

    eligible_result = await db.execute(
        select(func.count(Employee.id))
        .where(Employee.start_date <= ninety_days_ago)
        .where(Employee.is_active == True)
    )
    eligible_count = eligible_result.scalar() or 0

    # Build response results
    response_results = []
    for response in responses:
        employee = response.employee

        # Build employee result
        employee_result = EmployeeResult(
            id=employee.id,
            telegram_id=employee.telegram_id,
            telegram_username=employee.telegram_username,
            first_name=employee.first_name,
            last_name=employee.last_name,
        )

        # Build question results
        question_results = []
        for answer in response.answers:
            # Find question
            question = next((q for q in survey.questions if q.id == answer.question_id), None)
            if not question:
                continue

            # Get option texts for choice answers
            answer_options_texts = None
            if answer.answer_options:
                answer_options_texts = []
                for option_id in answer.answer_options:
                    option = next((o for o in question.options if o.id == option_id), None)
                    if option:
                        answer_options_texts.append(option.option_text)

            question_result = QuestionResult(
                question_id=answer.question_id,
                question_text=question.question_text,
                question_type=question.question_type,
                answer_text=answer.answer_text,
                answer_options=answer_options_texts,
            )
            question_results.append(question_result)

        response_result = ResponseResult(
            response_id=response.id,
            employee=employee_result,
            completed_at=response.completed_at,
            answers=question_results,
        )
        response_results.append(response_result)

    # Calculate completion rate
    completion_rate = 0.0
    if eligible_count > 0:
        completion_rate = len([r for r in response_results if r.completed_at]) / eligible_count

    return SurveyResults(
        survey_id=survey.id,
        survey_title=survey.title,
        responses=response_results,
        total_responses=len(response_results),
        completion_rate=completion_rate,
    )


@router.get("/{response_id}", response_model=SurveyResponseSchema)
async def get_response(response_id: int, db: AsyncSession = Depends(get_db)):
    """Get response by ID."""
    result = await db.execute(
        select(SurveyResponse)
        .where(SurveyResponse.id == response_id)
        .options(selectinload(SurveyResponse.answers))
    )
    response = result.scalar_one_or_none()

    if not response:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Response not found"
        )

    return response


@router.get("/surveys/{survey_id}/analytics", response_model=SurveyAnalytics)
async def get_survey_analytics(survey_id: int, db: AsyncSession = Depends(get_db)):
    """Get survey analytics with aggregated data for charts."""
    # Get survey with questions and options
    survey_result = await db.execute(
        select(Survey)
        .where(Survey.id == survey_id)
        .options(selectinload(Survey.questions).selectinload(Question.options))
    )
    survey = survey_result.scalar_one_or_none()

    if not survey:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Survey not found"
        )

    # Get all completed responses with answers
    responses_result = await db.execute(
        select(SurveyResponse)
        .where(SurveyResponse.survey_id == survey_id)
        .where(SurveyResponse.status == "completed")
        .options(selectinload(SurveyResponse.answers))
    )
    responses = responses_result.scalars().all()

    # Build question analytics
    question_analytics = []

    for question in survey.questions:
        # Get all answers for this question
        all_answers = []
        for response in responses:
            answer = next((a for a in response.answers if a.question_id == question.id), None)
            if answer:
                all_answers.append(answer)

        question_analytic = QuestionAnalytics(
            question_id=question.id,
            question_text=question.question_text,
            question_type=question.question_type,
            total_answers=len(all_answers),
        )

        if question.question_type in ("single_choice", "multiple_choice"):
            # Calculate distribution for choice questions
            option_counts = {option.id: 0 for option in question.options}

            for answer in all_answers:
                if answer.answer_options:
                    for option_id in answer.answer_options:
                        if option_id in option_counts:
                            option_counts[option_id] += 1

            # Build distribution list
            distribution = []
            for option in question.options:
                count = option_counts.get(option.id, 0)
                percentage = (count / len(all_answers) * 100) if all_answers else 0
                distribution.append({
                    "option_id": option.id,
                    "option": option.option_text,
                    "count": count,
                    "percentage": round(percentage, 2)
                })

            question_analytic.choice_distribution = distribution

        elif question.question_type == "text":
            # Collect all text responses
            text_responses = [
                answer.answer_text
                for answer in all_answers
                if answer.answer_text
            ]
            question_analytic.text_responses = text_responses

        question_analytics.append(question_analytic)

    # Calculate completion metrics
    total_responses = len(responses)
    completion_rate = 0.0

    # Get eligible employees count
    from datetime import timedelta
    ninety_days_ago = datetime.now().date() - timedelta(days=survey.days_after_start)

    eligible_result = await db.execute(
        select(func.count(Employee.id))
        .where(Employee.start_date <= ninety_days_ago)
        .where(Employee.is_active == True)
    )
    eligible_count = eligible_result.scalar() or 0

    if eligible_count > 0:
        completion_rate = total_responses / eligible_count

    return SurveyAnalytics(
        survey_id=survey.id,
        survey_title=survey.title,
        total_responses=total_responses,
        completed_responses=total_responses,
        completion_rate=completion_rate,
        question_analytics=question_analytics,
    )


@router.get("/employees/{employee_id}/responses", response_model=SurveyResults)
async def get_employee_responses(employee_id: int, db: AsyncSession = Depends(get_db)):
    """Get all survey responses for a specific employee."""
    # Get employee
    employee_result = await db.execute(
        select(Employee).where(Employee.id == employee_id)
    )
    employee = employee_result.scalar_one_or_none()

    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Employee not found"
        )

    # Get all responses for this employee with answers and survey
    responses_result = await db.execute(
        select(SurveyResponse)
        .where(SurveyResponse.employee_id == employee_id)
        .options(
            selectinload(SurveyResponse.answers),
            selectinload(SurveyResponse.survey).selectinload(Survey.questions).selectinload(Question.options)
        )
    )
    responses = responses_result.scalars().all()

    # Build response results for each survey
    response_results = []
    survey_ids = set()

    for response in responses:
        survey_ids.add(response.survey_id)

        # Build employee result
        employee_result = EmployeeResult(
            id=employee.id,
            telegram_id=employee.telegram_id,
            telegram_username=employee.telegram_username,
            first_name=employee.first_name,
            last_name=employee.last_name,
        )

        # Get survey for this response
        survey = response.survey

        # Build question results
        question_results = []
        for answer in response.answers:
            # Find question
            question = next((q for q in survey.questions if q.id == answer.question_id), None)
            if not question:
                continue

            # Get option texts for choice answers
            answer_options_texts = None
            if answer.answer_options:
                answer_options_texts = []
                for option_id in answer.answer_options:
                    option = next((o for o in question.options if o.id == option_id), None)
                    if option:
                        answer_options_texts.append(option.option_text)

            question_result = QuestionResult(
                question_id=answer.question_id,
                question_text=question.question_text,
                question_type=question.question_type,
                answer_text=answer.answer_text,
                answer_options=answer_options_texts,
            )
            question_results.append(question_result)

        response_result = ResponseResult(
            response_id=response.id,
            employee=employee_result,
            completed_at=response.completed_at,
            answers=question_results,
        )
        response_results.append(response_result)

    # Use first survey title if available
    survey_title = "Все опросы сотрудника"
    if response_results:
        first_response = await db.execute(
            select(SurveyResponse)
            .where(SurveyResponse.id == response_results[0].response_id)
            .options(selectinload(SurveyResponse.survey))
        )
        first_survey = first_response.scalar_one_or_none()
        if first_survey and first_survey.survey:
            survey_title = f"Опросы сотрудника: {employee.first_name} {employee.last_name}"

    return SurveyResults(
        survey_id=0,  # Multiple surveys, so 0
        survey_title=survey_title,
        responses=response_results,
        total_responses=len(response_results),
        completion_rate=1.0 if response_results else 0.0,
    )


@router.get("/surveys/{survey_id}/results/export")
async def export_survey_results_excel(survey_id: int, db: AsyncSession = Depends(get_db)):
    """Export survey results to Excel file."""
    # Get survey with questions and options
    survey_result = await db.execute(
        select(Survey)
        .where(Survey.id == survey_id)
        .options(selectinload(Survey.questions).selectinload(Question.options))
    )
    survey = survey_result.scalar_one_or_none()

    if not survey:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Survey not found"
        )

    # Get all responses with answers and employee
    responses_result = await db.execute(
        select(SurveyResponse)
        .where(SurveyResponse.survey_id == survey_id)
        .options(
            selectinload(SurveyResponse.answers),
            selectinload(SurveyResponse.employee)
        )
    )
    responses = responses_result.scalars().all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты опроса"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Create header row
    headers = ["№", "Сотрудник", "Telegram Username", "Дата завершения"]
    for question in survey.questions:
        headers.append(question.question_text)

    ws.append(headers)

    # Apply header style
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Add data rows
    for idx, response in enumerate(responses, 1):
        employee = response.employee
        completed_at = response.completed_at.strftime("%d.%m.%Y %H:%M") if response.completed_at else "Не завершено"

        row_data = [
            idx,
            f"{employee.last_name or ''} {employee.first_name or ''}".strip() or "-",
            employee.telegram_username or "-",
            completed_at
        ]

        # Add answers for each question
        for question in survey.questions:
            answer = next((a for a in response.answers if a.question_id == question.id), None)

            if answer:
                if question.question_type == "text":
                    row_data.append(answer.answer_text or "-")
                elif question.question_type in ("single_choice", "multiple_choice"):
                    if answer.answer_options:
                        option_texts = []
                        for option_id in answer.answer_options:
                            option = next((o for o in question.options if o.id == option_id), None)
                            if option:
                                option_texts.append(option.option_text)
                        row_data.append(", ".join(option_texts) if option_texts else "-")
                    else:
                        row_data.append("-")
                else:
                    row_data.append("-")
            else:
                row_data.append("-")

        ws.append(row_data)

        # Apply cell styles to data row
        for cell in ws[ws.max_row]:
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with survey title and date
    safe_title = "".join(c for c in survey.title if c.isalnum() or c in (' ', '-', '_')).strip()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"survey_{survey.id}_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )
